import openpyxl
import os
from operator import itemgetter
from itertools import groupby


os.chdir('D:\\GitHub\\compareExTables')
# ИМЯ ФАЙЛА
wb1 = openpyxl.load_workbook('comparison 12 09 2018_4.xlsx', data_only=True)
wb2 = openpyxl.load_workbook('deficit BB.xlsx', data_only=True)

noneList = (None, ' ', '')
# ЛИСТЫ
resultSheet = wb1.get_sheet_by_name('Result')
defSheet = wb2.get_sheet_by_name('all')
defResult = wb2.create_sheet('all_result')

# ДИАПАЗОНЫ В ЛИСТАХ
resRange = range(1, 11198)
defRange = range(3, 11418)

# сбор данных для сравнения
# колонка О в overall, колонки С и I в LISAP
resID_Count = [[resultSheet.cell(row = i, column = 1).value, resultSheet.cell(row = i, column = 2).value] for i in resRange]

defID = [defSheet.cell(row = i, column = 15).value for i in defRange if defSheet.cell(row = i, column = 15).value not in noneList]

out = []
for i in defID:
	a = list(filter(lambda x: x[0] == i, resID_Count))
	if a:
		out.append([i,'deficit BB', a[0][1]])
	else:
		out.append([None, None, None])


r = 1
for i in out:
	defResult.cell(row = r, column = 1).value = i[0]
	defResult.cell(row = r, column = 2).value = i[1]
	defResult.cell(row = r, column = 3).value = i[2]
	r += 1

wb2.save('deficit BB_6.xlsx')

print("out")
