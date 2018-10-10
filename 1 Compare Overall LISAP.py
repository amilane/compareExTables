import openpyxl
import os
from operator import itemgetter
from itertools import groupby


os.chdir('D:\\GitHub\\compareExTables')
# ИМЯ ФАЙЛА
filename = 'comparison  09 2018'

wb = openpyxl.load_workbook(filename + '.xlsx')

# ЛИСТЫ
overSheet = wb.get_sheet_by_name('overall list of details2')
lisapSheet = wb.get_sheet_by_name('LISAP')
resultSheet = wb.create_sheet('Result')
addToOverall = wb.create_sheet('Add To Overall')
notInOverSheet = wb.create_sheet('Not in Overall')

# ДИАПАЗОНЫ В ЛИСТАХ
overRange = range(4, 11285)
lisapRange = range(2, 70081)

# сбор данных для сравнения
# колонка О в overall, колонки С и I в LISAP
overID_Count = [[overSheet.cell(row = i, column = 10).value, overSheet.cell(row = i, column = 15).value] for i in overRange]
_overCount = [i[0] for i in overID_Count]
overCount = []
for i in _overCount:
	if i == None or str(i).strip() == "":
		overCount.append(0)
	else:
		overCount.append(i)

overID = [i[1] for i in overID_Count]

lisID_NUM = [[lisapSheet.cell(row = i, column = 3).value, lisapSheet.cell(row = i, column = 9).value] \
				for i in lisapRange \
				if str(lisapSheet.cell(row = i, column = 3).value).strip() != '' \
				and str(lisapSheet.cell(row = i, column = 9).value).strip() != '']

lis_ID = [i[0] for i in lisID_NUM]
lis_NUM = [i[1] for i in lisID_NUM]

lisID_in = []
for i, n in zip(overID, overCount):
	n = int(n)
	if i in lis_ID:
		numbers = ''
		while n > 0 and i in lis_ID:
			ind = lis_ID.index(i)
			numbers += lis_NUM[ind] + ', '
			lis_ID.pop(ind)
			lis_NUM.pop(ind)
			n -= 1
		lisID_in.append(numbers)
	else:
		lisID_in.append('')

dop_ID_NUM = list(filter(lambda i: i[0] in overID, list(zip(lis_ID, lis_NUM))))
lisID_out = list(filter(lambda i: i[0] not in overID, list(zip(lis_ID, lis_NUM))))

for i, k in zip(overRange, lisID_in):
  resultSheet.cell(row = i, column = 21).value = k

for i in range(1, len(lisID_out)+1):
  notInOverSheet.cell(row = i, column = 1).value = lisID_out[i-1][0]
  notInOverSheet.cell(row = i, column = 2).value = lisID_out[i-1][1]

for i in range(1, len(dop_ID_NUM)+1):
  addToOverall.cell(row = i, column = 15).value = dop_ID_NUM[i-1][0]
  addToOverall.cell(row = i, column = 21).value = dop_ID_NUM[i-1][1]

wb.save(filename + '_2.xlsx')

print("OK")



