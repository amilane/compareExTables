import openpyxl
import os
from operator import itemgetter
from itertools import groupby


os.chdir('D:\\GitHub\\compareExTables')
# ИМЯ ФАЙЛА
filename = 'comparison 12 09 2018'

wb = openpyxl.load_workbook(filename + '.xlsx', data_only=True)

noneList = (None, ' ', '')
# ЛИСТЫ
resultOfAnalysisSheet = wb.get_sheet_by_name('RESULTS OF ANALYSIS')
lisapSheet = wb.get_sheet_by_name('LISAP')
resultSheet = wb.create_sheet('Result')

# ДИАПАЗОНЫ В ЛИСТАХ
resRange = range(4, 11151)
lisapRange = range(2, 70080)

# сбор данных для сравнения
# колонка О в overall, колонки С и I в LISAP
resID_Count = [[resultOfAnalysisSheet.cell(row = i, column = 1).value, resultOfAnalysisSheet.cell(row = i, column = 2).value] for i in resRange]
_resCount = [i[1] for i in resID_Count]
resCount = []
for i in _resCount:
	if i in noneList:
		resCount.append(0)
	else:
		resCount.append(int(i))

resID = [i[0] for i in resID_Count]

lisID_NUM = [[lisapSheet.cell(row = i, column = 3).value, lisapSheet.cell(row = i, column = 9).value] \
				for i in lisapRange \
				if lisapSheet.cell(row = i, column = 3).value not in noneList \
				and lisapSheet.cell(row = i, column = 9).value not in noneList]

lis_ID = [i[0] for i in lisID_NUM]
lis_NUM = [i[1] for i in lisID_NUM]

lisID_in = []
for i, n in zip(resID, resCount):
	if i in lis_ID:
		numbers = []
		while n > 0 and i in lis_ID:
			ind = lis_ID.index(i)
			numbers.append(lis_NUM[ind])
			lis_ID.pop(ind)
			lis_NUM.pop(ind)
			n -= 1
		numbers.sort()
		grNumbers = [[x for x in g] for k,g in groupby(numbers)]
		strNum = ''
		for g in grNumbers:
			if len(g) > 1:
				strNum += f"{g[0]}({len(g)}), "
			elif len(g) == 1:
				strNum += f"{g[0]}, "
			else:
				strNum = ''
		lisID_in.append(strNum)
	else:
		lisID_in.append('')

for i, k in zip(resRange, lisID_in):
  resultSheet.cell(row = i, column = 1).value = k

wb.save(filename + '_3.xlsx')

print("OK")



