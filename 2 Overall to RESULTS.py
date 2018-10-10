import openpyxl
import os
from operator import itemgetter
from itertools import groupby


os.chdir('D:\\GitHub\\compareExTables')
# ИМЯ ФАЙЛА
filename = 'comparison 12 09 2018'

wb = openpyxl.load_workbook(filename + '.xlsx')

noneList = (None, ' ', '')
# ЛИСТЫ
overSheet = wb.get_sheet_by_name('overall list of details2')
resultOfAnalysisSheet = wb.get_sheet_by_name('RESULTS OF ANALYSIS')
notInResultOfAnalysis = wb.create_sheet('Not in ResultOfAnalysis')

# ДИАПАЗОНЫ В ЛИСТАХ
overRange = range(4, 11932)
resOfAnRange = range(4, 9728)

# сбор данных для сравнения
# колонка О в overall, колонка А в RESULTS OF ANALYSIS
overID = [overSheet.cell(row = i, column = 15).value for i in overRange if overSheet.cell(row = i, column = 15).value not in noneList]
resID = [resultOfAnalysisSheet.cell(row = i, column = 1).value for i in resOfAnRange]
notInResID = [i for i in overID if i not in resID]

for i in range(1, len(notInResID)+1):
  notInResultOfAnalysis.cell(row = i, column = 1).value = notInResID[i-1]

wb.save(filename + '_2.xlsx')	

print("OK")

