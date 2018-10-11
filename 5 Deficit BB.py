import openpyxl
import os
from operator import itemgetter
from itertools import groupby


os.chdir('D:\\GitHub\\compareExTables')
# ИМЯ ФАЙЛА
filename = 'comparison 12 09 2018_4'

wb1 = openpyxl.load_workbook(filename + '.xlsx', data_only=True)
wb2 = openpyxl.load_workbook('deficit BB.xlsx', data_only=True)

noneList = (None, ' ', '')
# ЛИСТЫ
resultSheet = wb1.get_sheet_by_name('Result')
defSheet = wb2.get_sheet_by_name('all')

# ДИАПАЗОНЫ В ЛИСТАХ
resRange = range(1, 11198)
defRange = range(3, 11418)

# сбор данных для сравнения
# колонка О в overall, колонки С и I в LISAP
resID = [resultSheet.cell(row = i, column = 1).value for i in resRange]

defID_Count = [[defSheet.cell(row = i, column = 15).value, defSheet.cell(row = i, column = 10).value] for i in defRange\
		if defSheet.cell(row = i, column = 15).value not in noneList]
for i in defID_Count:
	if i[1] in noneList:
		i[1] = 0


defID_Count.sort(key=lambda i: i[0])

g_defID_Count = [[x for x in g] for k,g in groupby(defID_Count, lambda i: i[0])]

defID_Count = []
for g in g_defID_Count:
	if len(g) > 1:
		s = sum([i[1] for i in g])
		defID_Count.append([g[0][0], s])
	else:
		defID_Count.append(g[0])

out = []
for i in resID:
	a = list(filter(lambda x: x[0] == i, defID_Count))
	if a:
		out.append(['dedicit BB', a[0][1]])
	else:
		out.append([None, None])





r = 1
for i in out:
	resultSheet.cell(row = r, column = 7).value = i[0]
	resultSheet.cell(row = r, column = 8).value = i[1]
	r += 1

wb1.save(filename + '_5.xlsx')

print("out")
