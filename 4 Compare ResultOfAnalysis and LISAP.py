import openpyxl
import os
from operator import itemgetter
from itertools import groupby


os.chdir('D:\\GitHub\\compareExTables')
# ИМЯ ФАЙЛА
filename = 'comparison 12 09 2018'

wb = openpyxl.load_workbook(filename + '.xlsx', data_only=True)

noneList = (None, ' ', '')
# ЛИСТЫ
overallSheet = wb.get_sheet_by_name('overall list of details2')
resultOfAnalysisSheet = wb.get_sheet_by_name('RESULTS OF ANALYSIS')
lisapSheet = wb.get_sheet_by_name('LISAP')
resultSheet = wb.create_sheet('Result')

# ДИАПАЗОНЫ В ЛИСТАХ
overRange = range(4, 11932)
resRange = range(4, 11151)
lisapRange = range(2, 70080)

# сбор данных для сравнения
# колонка О в overall, колонки С и I в LISAP
resID_Count = [[resultOfAnalysisSheet.cell(row = i, column = 1).value,\
		 resultOfAnalysisSheet.cell(row = i, column = 2).value]\
		  for i in resRange\
			if resultOfAnalysisSheet.cell(row = i, column = 1).value not in noneList\
			and resultOfAnalysisSheet.cell(row = i, column = 2).value not in noneList]

res_Comments = [[resultOfAnalysisSheet.cell(row = i, column = 1).value,\
		 resultOfAnalysisSheet.cell(row = i, column = 8).value]\
		  for i in resRange\
			if resultOfAnalysisSheet.cell(row = i, column = 1).value not in noneList\
			and resultOfAnalysisSheet.cell(row = i, column = 8).value not in noneList]
diComments = dict(res_Comments)

resID_Count.sort(key=lambda i: i[0])
g_resID_Count = [[x for x in g] for k,g in groupby(resID_Count, lambda i: i[0])]

resID_Count = []
for g in g_resID_Count:
	if len(g) > 1:
		s = sum([i[1] for i in g])
		resID_Count.append([g[0][0], s])
	else:
		resID_Count.append(g[0])

overID_Count = [[overallSheet.cell(row = i, column = 15).value,\
		 overallSheet.cell(row = i, column = 10).value]\
		  for i in overRange\
			if overallSheet.cell(row = i, column = 15).value not in noneList]
for i in overID_Count:
	if i[1] in noneList:
		i[1] = 0

overID_Count.sort(key=lambda i: i[0])
g_overID_Count = [[x for x in g] for k,g in groupby(overID_Count, lambda i: i[0])]

overID_Count = []
for g in g_overID_Count:
	if len(g) > 1:
		s = sum([i[1] for i in g])
		overID_Count.append([g[0][0], s])
	else:
		overID_Count.append(g[0])

Common = overID_Count.copy()
for i in resID_Count:
	if i not in Common:
		Common.append(i)


lisID_NUM = [[lisapSheet.cell(row = i, column = 3).value, lisapSheet.cell(row = i, column = 9).value] \
				for i in lisapRange \
				if lisapSheet.cell(row = i, column = 3).value not in noneList \
				and lisapSheet.cell(row = i, column = 9).value not in noneList]

for i in Common:
	a = list(filter(lambda x: x[0] == i[0], lisID_NUM))
	if a:
		nums = [i[1] for i in a]
		nums.sort()
		grNums = [[x for x in g] for k,g in groupby(nums)]
		strNum = ''
		for g in grNums:
			if len(g) > 1:
				strNum += f"{g[0]}({len(g)}), "
			elif len(g) == 1:
				strNum += f"{g[0]}, "
			else:
				strNum = None
	else:
		strNum = None
	over = None
	res = None
	if i in overID_Count:
		over = 'overall'
	if i in resID_Count:
		res = 'result of analysis'
	i.append(strNum)
	i.append(over)
	i.append(res)
	if i[0] in diComments:
		i.append(diComments[i[0]])
	else:
		i.append(None)

r = 1
for i in Common:
	resultSheet.cell(row = r, column = 1).value = i[0]
	resultSheet.cell(row = r, column = 2).value = i[1]
	resultSheet.cell(row = r, column = 3).value = i[2]
	resultSheet.cell(row = r, column = 4).value = i[3]
	resultSheet.cell(row = r, column = 5).value = i[4]
	resultSheet.cell(row = r, column = 6).value = i[5]
	r += 1

wb.save(filename + '_4.xlsx')

print("OK")



